import openpyxl
import json

sheet = openpyxl.load_workbook("./AT17-ShipPoints.xlsx", data_only=True).active
pointsDict = []
nameCol, pointsCol, hullTypeCol = 5, 7, 8

for row in range(3, 243):
  name = sheet.cell(row, column = nameCol).value
  points = sheet.cell(row, column = pointsCol).value
  hullType = sheet.cell(row, column = hullTypeCol).value

  pointsDict.append({ 'shipId': row-2,'name': name, 'points': points, 'hullType': hullType, 'fitting': ''})

with open('../public/DO_NOT_EDIT-AT17-ShipPoints.json', 'w', encoding='utf-8') as f:
  json.dump(pointsDict, f, indent=2)